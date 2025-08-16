#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件预览工具

功能：查看Excel文件的结构和J列内容，用于了解数据格式
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def preview_excel_file(file_path):
    """
    预览Excel文件内容
    
    Args:
        file_path: Excel文件路径
    """
    print(f"📖 正在预览文件：{file_path}")
    
    try:
        # 打开工作簿
        workbook = openpyxl.load_workbook(file_path)
        print(f"📊 工作簿包含工作表：{workbook.sheetnames}")
        
        # 获取活动工作表
        worksheet = workbook.active
        print(f"📋 当前工作表：{worksheet.title}")
        print(f"📏 工作表尺寸：{worksheet.max_row} 行 x {worksheet.max_column} 列")
        
        # 显示表头（第一行）
        print("\n📑 表头信息：")
        for col in range(1, min(worksheet.max_column + 1, 15)):  # 最多显示前15列
            header_cell = worksheet.cell(row=1, column=col)
            column_letter = get_column_letter(col)
            print(f"  {column_letter}列: {header_cell.value}")
        
        # 重点查看J列（第10列）
        j_column = 10
        column_letter = get_column_letter(j_column)
        print(f"\n🔍 {column_letter}列（第{j_column}列）前10行数据：")
        print("-" * 60)
        
        for row in range(1, min(11, worksheet.max_row + 1)):
            cell = worksheet.cell(row=row, column=j_column)
            cell_value = cell.value
            
            # 显示行号、值和数据类型
            if cell_value is None:
                print(f"  第{row:2d}行: (空)")
            else:
                value_str = str(cell_value)
                if len(value_str) > 50:
                    value_str = value_str[:47] + "..."
                print(f"  第{row:2d}行: {value_str}")
        
        # 统计J列的数据类型
        print(f"\n📊 {column_letter}列数据统计：")
        url_count = 0
        empty_count = 0
        text_count = 0
        
        for row in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=j_column)
            if cell.value is None or cell.value == "":
                empty_count += 1
            elif str(cell.value).strip().startswith(('http://', 'https://')):
                url_count += 1
            else:
                text_count += 1
        
        print(f"  总行数: {worksheet.max_row}")
        print(f"  URL格式: {url_count}")
        print(f"  普通文本: {text_count}")
        print(f"  空白单元格: {empty_count}")
        
    except Exception as e:
        print(f"❌ 预览过程中发生错误：{str(e)}")

def main():
    """主函数"""
    file_path = "projects/AI学术分析/智能整理/data/processed/Hinton_with_Affiliations_v5.xlsx"
    
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        return
    
    print("🔍 Excel文件预览")
    print("=" * 50)
    preview_excel_file(file_path)

if __name__ == "__main__":
    main() 