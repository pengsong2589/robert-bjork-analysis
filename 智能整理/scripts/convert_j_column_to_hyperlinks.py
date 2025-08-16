#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel J列超链接转换工具

功能：将Excel文件中J列的本地文件路径转换为可点击的超链接
作者：AI助手
创建时间：2025年1月
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil
from datetime import datetime
from urllib.parse import urljoin
from pathlib import Path

def create_backup(file_path, backup_dir):
    """
    创建文件备份
    
    Args:
        file_path: 原文件路径
        backup_dir: 备份目录
    
    Returns:
        backup_path: 备份文件路径
    """
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    # 获取文件名和扩展名
    file_name = os.path.basename(file_path)
    name, ext = os.path.splitext(file_name)
    
    # 创建带时间戳的备份文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"backup_{timestamp}_{name}{ext}"
    backup_path = os.path.join(backup_dir, backup_name)
    
    # 复制文件
    shutil.copy2(file_path, backup_path)
    print(f"✅ 已创建备份文件：{backup_path}")
    
    return backup_path

def is_file_path(text):
    """
    判断文本是否为文件路径
    
    Args:
        text: 要检查的文本
    
    Returns:
        bool: 是否为文件路径
    """
    if not isinstance(text, str):
        return False
    
    text = text.strip()
    
    # 检查是否为绝对路径或相对路径
    if text.startswith('/') or text.startswith('\\') or (len(text) > 1 and text[1] == ':'):
        return True
    
    # 检查是否包含文件扩展名
    if '.' in text and len(text.split('.')[-1]) <= 5:
        return True
    
    return False

def path_to_file_url(file_path):
    """
    将本地文件路径转换为file:// URL
    
    Args:
        file_path: 本地文件路径
    
    Returns:
        str: 文件URL
    """
    # 使用pathlib处理路径
    path = Path(file_path)
    
    # 转换为file:// URL格式
    if os.name == 'nt':  # Windows系统
        # Windows路径需要特殊处理
        file_url = path.as_uri()
    else:  # Unix-like系统（macOS、Linux）
        file_url = path.as_uri()
    
    return file_url

def convert_j_column_to_hyperlinks(file_path):
    """
    将Excel文件J列的文件路径转换为超链接
    
    Args:
        file_path: Excel文件路径
    """
    print(f"📖 正在处理文件：{file_path}")
    
    # 1. 创建备份
    backup_dir = os.path.join(os.path.dirname(file_path), "..", "backup")
    backup_path = create_backup(file_path, backup_dir)
    
    try:
        # 2. 打开工作簿
        workbook = openpyxl.load_workbook(file_path)
        print(f"📊 工作簿已加载，包含工作表：{workbook.sheetnames}")
        
        # 3. 获取活动工作表（或第一个工作表）
        worksheet = workbook.active
        print(f"📋 正在处理工作表：{worksheet.title}")
        
        # 4. 获取J列（第10列）的数据
        j_column = 10
        column_letter = get_column_letter(j_column)
        print(f"🔍 正在处理 {column_letter} 列（{worksheet.cell(row=1, column=j_column).value}）...")
        
        # 5. 统计信息
        total_rows = 0
        converted_count = 0
        empty_count = 0
        invalid_count = 0
        file_not_exist_count = 0
        
        # 6. 遍历J列的所有行（跳过表头）
        for row in range(2, worksheet.max_row + 1):  # 从第2行开始，跳过表头
            cell = worksheet.cell(row=row, column=j_column)
            total_rows += 1
            
            if cell.value is None or cell.value == "":
                empty_count += 1
                continue
            
            cell_value = str(cell.value).strip()
            
            # 检查是否为有效的文件路径
            if is_file_path(cell_value):
                # 检查文件是否存在
                if os.path.exists(cell_value):
                    try:
                        # 转换为file:// URL
                        file_url = path_to_file_url(cell_value)
                        
                        # 创建超链接
                        cell.hyperlink = file_url
                        cell.style = "Hyperlink"  # 应用超链接样式
                        
                        # 可选：修改显示文本为文件名
                        # cell.value = os.path.basename(cell_value)
                        
                        converted_count += 1
                        print(f"  ✅ 第{row}行：{os.path.basename(cell_value)}")
                    except Exception as e:
                        print(f"  ❌ 第{row}行：转换失败 - {str(e)}")
                        invalid_count += 1
                else:
                    file_not_exist_count += 1
                    print(f"  ⚠️  第{row}行：文件不存在 - {cell_value[:50]}...")
            else:
                invalid_count += 1
                print(f"  ⚠️  第{row}行：非文件路径格式 - {cell_value[:30]}...")
        
        # 7. 保存文件
        workbook.save(file_path)
        print(f"💾 文件已保存：{file_path}")
        
        # 8. 输出统计信息
        print("\n📊 处理统计：")
        print(f"  总数据行数：{total_rows} (不包括表头)")
        print(f"  成功转换：{converted_count}")
        print(f"  空白单元格：{empty_count}")
        print(f"  文件不存在：{file_not_exist_count}")
        print(f"  无效路径：{invalid_count}")
        
        if converted_count > 0:
            print(f"\n✅ 成功将 {converted_count} 个文件路径转换为超链接！")
            print("💡 点击超链接将会打开对应的PDF文件")
        else:
            print("\n⚠️  没有找到有效的文件路径进行转换")
        
    except Exception as e:
        print(f"❌ 处理过程中发生错误：{str(e)}")
        print(f"💡 可以从备份文件恢复：{backup_path}")
        raise

def main():
    """主函数"""
    # Excel文件路径
    file_path = "projects/AI学术分析/智能整理/data/processed/Hinton_with_Affiliations_v5.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在：{file_path}")
        return
    
    print("🚀 开始Excel J列文件路径超链接转换...")
    print("=" * 60)
    
    try:
        convert_j_column_to_hyperlinks(file_path)
        print("\n🎉 转换完成！")
        print("\n📝 使用说明：")
        print("1. 打开Excel文件")
        print("2. 点击J列中的蓝色超链接文本")
        print("3. 系统将自动打开对应的PDF文件")
    except Exception as e:
        print(f"\n❌ 转换失败：{str(e)}")

if __name__ == "__main__":
    main() 